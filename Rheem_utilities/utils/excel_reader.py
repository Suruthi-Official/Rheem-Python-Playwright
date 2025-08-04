import openpyxl
from collections import defaultdict

class ExcelReader:
    """
    Utility class to read data from Excel files.
    This class uses openpyxl to interact with Excel sheets and retrieve data in a structured format.
    """

    def __init__(self):
        # Nested dict to store objects from the Excel file: SheetName -> Variable -> (Locator -> Value)
        self.objects = defaultdict(lambda: defaultdict(dict))

    def get_page_objects(self, excel_file_path):
        """
        Reads page objects from an Excel file and organizes them into a structured dict.
        :param excel_file_path: The file path of the Excel document.
        :return: A dict containing the structured page objects from the Excel file.
        """
        sheet_names = self.get_sheet_names(excel_file_path)
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            for page in sheet_names:
                sheet = workbook[page]
                variables = defaultdict(dict)
                # Assumes first row is header: variables, locator, value
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                var_idx = headers.index('variables')
                locator_idx = headers.index('locator')
                value_idx = headers.index('value')
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    variable = row[var_idx]
                    locator = row[locator_idx]
                    value = row[value_idx]
                    if variable and locator:
                        variables[variable][locator] = value
                self.objects[page] = variables
        except Exception as e:
            print(e)
        return dict(self.objects)

    def get_sheet_names(self, excel_file_path):
        """
        Retrieves a list of sheet names from the specified Excel file.
        :param excel_file_path: The path to the Excel file from which to retrieve the sheet names.
        :return: A list of strings containing the names of all sheets in the Excel file.
        """
        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            print(e)
            return []
